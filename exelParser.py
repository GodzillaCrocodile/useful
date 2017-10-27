# 'EStroev'
import os
import openpyxl
import re
import argparse
from datetime import datetime

inputFileCSV = r"D:\Working\Projects\Network\Regions orig.csv"
inputFileXLSX = r"D:\Working\Projects\Network\Regions orig.xlsx"
outFile = r"D:\Working\Projects\Network\Regions_parse.xlsx"

ip_pattern = re.compile('(\d+)\.(\d+)\.(\d+)\.(\d+)/\d+')

def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]


def row_splitter(row):
    for line in row.split('\n'):
        ip, description = None, None
        net = line.strip('_x000D_')
        if len(net.split(' - ')) > 1:
            ip, description = net.split(' - ')
        else:
            ip = net
        if ip:
            if not ip[0].isdigit():
                ip = None
            else:
                ip_re = ip_pattern.search(ip)
                if ip_re:
                    ip = ip_re.group()

        if description:
            description = description.strip()

        yield [ip, description]


def parser(ws):
    data = list()
    title = [cell.value for cell in ws[1]]
    rows = list(iter_rows(ws))
    for row in list(rows[1:]):
        if row:
            name, network_office_1, network_office_2 = row
            if network_office_1:
                for net in row_splitter(network_office_1):
                    ip_office_1, description_office_1 = net
                    data.append([name, ip_office_1, description_office_1])
            if network_office_2:
                for net in row_splitter(network_office_2):
                    ip_office_2, desciption_office_2 = net
                    data.append([name, ip_office_2, desciption_office_2])

    return data


def write_data(outPath, title, data):
    if not os.path.exists(outPath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
    else:
        wb = openpyxl.load_workbook(outPath)
        ws = wb.create_sheet(title)

    for i in range(len(data[1:])):
        for j in range(len(data[0])):
            ws.cell(row=i+2, column=j+1).value = data[i+1][j]

    wb.save(outPath)
    print('[+] Write "{}" ({} entries) to {}'.format(title, len(data), outPath))


def main():


    startTime = datetime.now()
    print(startTime.strftime('[*] Start time: %d.%m.%Y %H:%M:%S'))

    inWB = openpyxl.load_workbook(filename=inputFileXLSX)
    print(f'[+] Open {inputFileXLSX}')
    data = parser(inWB['Лист1'])
    write_data(outFile, 'Test', data)

    endTime = datetime.now()
    print('[*] Total elapsed time - {0} seconds'.format((endTime - startTime).seconds))
    print(endTime.strftime('[*] End time: %d.%m.%Y %H:%M:%S'))

if __name__ == '__main__':
    main()