# 'EStroev'
import os
import openpyxl
import re
import argparse
from datetime import datetime

ip_pattern = re.compile('(\d+)\.(\d+)\.(\d+)\.(\d+)/\d+')


def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]


def row_splitter(row):
    for line in row.split('\n'):
        ip, description = None, None
        net = line.strip('_x000D_')
        if len(net.split(',')) > 1:
            ip = net.split(',')
        elif len(net.split(' - ')) > 1:
            ip, description = net.split(' - ')
        else:
            ip = net
        if ip:
            if type(ip) == str:
                if not ip[0].isdigit():
                    ip = None
                elif len(net.split(' - ')) > 1:
                    ip, description = net.split(' - ')
                else:
                    ip_re = ip_pattern.search(ip)
                    if ip_re:
                        ip = ip_re.group()

                if description:
                    description = description.strip()
                if ip:
                    ip = ip.strip()

                yield [ip, description]
            elif type(ip) == list:
                for ipIter in ip:
                    if ipIter:
                        if len(ipIter.split(' - ')) > 1:
                            ipIter, description = ipIter.split(' - ')
                            if description:
                                description = description.strip()
                            if ipIter:
                                ipIter = ipIter.strip()

                            yield [ipIter, description]
                        else:
                            if description:
                                description = description.strip()
                            if ipIter:
                                ipIter = ipIter.strip()

                            yield [ipIter, description]


def worker(ws):
    data = list()
    title = [cell.value for cell in ws[1]]
    rows = list(iter_rows(ws))
    for row in list(rows[1:]):
        if row:
            name, networkOffice1, networkOffice2 = row
            if networkOffice1:
                for net in row_splitter(networkOffice1):
                    ipOffice1, descriptionOffice1 = net
                    data.append([name, ipOffice1, descriptionOffice1])
            if networkOffice2:
                for net in row_splitter(networkOffice2):
                    ipOffice2, desciptionOffice2 = net
                    data.append([name, ipOffice2, desciptionOffice2])

    return data


def worker2(ws):
    data = list()
    title = [cell.value for cell in ws[1]]
    rows = list(iter_rows(ws))
    for row in list(rows[1:]):
        if row:
            name, network, timeDelta = row
            if network:
                for net in row_splitter(network):
                    ip, description = net
                    data.append([name, ip, description])
    return data


def write_data(outPath, title, data):
    if not os.path.exists(outPath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
    else:
        wb = openpyxl.load_workbook(outPath)
        ws = wb.create_sheet(title)

    for i in range(len(data[1:])):
        for j in range(len(data[0])):
            ws.cell(row=i+2, column=j+1).value = data[i+1][j]

    wb.save(outPath)
    print('[+] Write "{}" ({} entries) to {}'.format(title, len(data), outPath))


def main():
    parser = argparse.ArgumentParser(description='Excel parser')

    parser.add_argument('-o', dest='outFile', action='store', help='Output file')
    parser.add_argument('-f', dest='inFile', action='store', help='Input file')

    args = parser.parse_args()
    if not args.inFile:
        print('[-] You must specify an existing path to the input file!')
        exit(-1)
    elif not os.path.exists(args.inFile):
        print('[-] Input file %s does not exist!' % os.path.abspath(args.inFile))
        exit(-1)
    if not args.outFile:
        print('[-] You must specify an existing path to the output folder!')
        exit(-1)

    startTime = datetime.now()
    print(startTime.strftime('[*] Start time: %d.%m.%Y %H:%M:%S'))

    inWB = openpyxl.load_workbook(filename=args.inFile)

    print(f'[+] Open {args.inFile}')
    data = worker2(inWB['Лист1'])
    write_data(args.outFile, 'Test', data)

    endTime = datetime.now()
    print('[*] Total elapsed time - {0} seconds'.format((endTime - startTime).seconds))
    print(endTime.strftime('[*] End time: %d.%m.%Y %H:%M:%S'))

if __name__ == '__main__':
    main()